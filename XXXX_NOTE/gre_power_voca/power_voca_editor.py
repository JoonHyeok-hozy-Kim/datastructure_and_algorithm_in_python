import openpyxl

power_voca_list = []
power_voce_raw = open(
    "C:/Users/Hubris Ozymandias/PycharmProjects/datastructure_and_algorithm_in_python/XXXX_NOTE/gre_power_voca/power_voca_raw.txt",
    'r',
    encoding='UTF8'
)
while True:
    line = power_voce_raw.readline()
    if not line:
        break
    else:
        enter_effaced = line.split('\n')[0]
        power_voca_list.append(enter_effaced.split('\t'))

# idx = 1
# for i in power_voca_list:
#     try:
#         if idx != int(i[0]):
#             print(i)
#             break
#         else:
#             idx += 1
#     except:
#         print('Error at : {}'.format(i))

# print(len(power_voca_list))
# for i in power_voca_list:
#     print(power_voca_list)

wb = openpyxl.Workbook()
ws = wb.active
wb.create_sheet('power_voca', 0)
for i in range(len(power_voca_list)):
    if int(power_voca_list[i][0])%120 == 1:
        chapter_name_list = ['Ch ']
        chapter_name_list.append(power_voca_list[i][0])
        chapter_name_list.append(' ~ ')
        chapter_name_list.append(str(int(power_voca_list[i][0])+119))
        ws.cell(row=i+1, column=1).value = ''.join(chapter_name_list)
    try:
        ws.cell(row=i+1, column=2).value = power_voca_list[i][1]
        if power_voca_list[i][2] == '':
            ws.cell(row=i+1, column=3).value = power_voca_list[i][3]
        else:
            ws.cell(row=i+1, column=3).value = power_voca_list[i][2]
    except:
        print('Error : {}'.format(power_voca_list[i]))
wb.save(filename='power_voca_temp.xlsx')