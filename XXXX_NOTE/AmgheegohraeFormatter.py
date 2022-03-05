import openpyxl

class AmgheegohraeFormat:

    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.word_dict_list = []

    def add_word_dict(self, word_dict):
        if len(word_dict) == 2:
            self.word_dict_list.append(word_dict)
        else:
            raise ValueError("add_word_dict Format error : word_dict format should be {'chapter_name', 'word_list'}")

    def __str__(self):
        result_list = []
        for word_dict in self.word_dict_list:
            result_list.append('Chapter : ')
            result_list.append(word_dict['chapter_name'])
            result_list.append('\n')
            for word in word_dict['word_list']:
                result_list.append('\t')
                result_list.append(word[0])
                result_list.append(' : ')
                result_list.append(word[1])
                result_list.append('\n')
        return ''.join(result_list)

    def save_in_excel(self, excel_file_name=None):
        if excel_file_name is None:
            excel_file_name = 'gorae'
        excel_file_name += '.xlsx'
        row_idx = 1
        # Null space required for the 1st row

        row_idx += 1
        self.ws.cell(row=row_idx, column=1).value = '챕터 이름'
        self.ws.cell(row=row_idx, column=2).value = '단어'
        self.ws.cell(row=row_idx, column=3).value = '뜻'
        self.ws.cell(row=row_idx, column=4).value = '발음(옵션)'
        self.ws.cell(row=row_idx, column=5).value = '예문(옵션)'
        self.ws.cell(row=row_idx, column=6).value = '예문 뜻(옵션)'

        row_idx += 1

        for word_dict in self.word_dict_list:
            self.ws.cell(row=row_idx, column=1).value = word_dict['chapter_name']
            for word in word_dict['word_list']:
                self.ws.cell(row=row_idx, column=2).value = word[0]
                self.ws.cell(row=row_idx, column=3).value = word[1]
                row_idx += 1

        self.wb.save(filename=excel_file_name)